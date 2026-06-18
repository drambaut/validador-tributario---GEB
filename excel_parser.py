from pathlib import Path
from openpyxl import load_workbook


def _admin(ws) -> dict:
    return {str(ws.cell(r, 1).value).strip(): ws.cell(r, 2).value for r in range(1, 20) if ws.cell(r, 1).value}


def parse_thomson(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        admin_ws = wb["Admin"] if "Admin" in wb.sheetnames else wb["ADMIN"]
        admin = _admin(admin_ws)
        out = {
        "_source": path.name,
        "nit": admin.get("EntityID"), "razon_social": admin.get("EntityName"),
        "anio": admin.get("Year"), "periodo": admin.get("Period"),
        "municipio": str(admin.get("Jurisdiction") or "").split("-")[0],
        "tipo": admin.get("TaxType"),
        }
        form_name = next((s for s in wb.sheetnames if "formulario" in s.lower()), None)
        if not form_name:
            return out
        ws = wb[form_name]
        # Formularios ICA/AutoICA nacionales de Ciénaga.
        if form_name.lower() == "formulario":
            cells = {
            "direccion": "K10", "telefono": "E13", "email": "J13",
            "ingresos_totales": "AB14", "ingresos_fuera": "AB15", "ingresos_municipio": "AB16",
            "base_gravable": "AB22", "impuesto_ica": "AB30", "avisos_tableros": "AB31",
            "sobretasa_bomberil": "AB33", "sobretasa_seguridad": "AB34", "total_impuesto_cargo": "AB35",
            "autorretenciones": "AB38", "valor_pagar": "AB48", "representante": "F56",
            "representante_id": "K57", "revisor": "U56", "revisor_id": "X57",
            }
            out.update({k: ws[ref].value for k, ref in cells.items()})
            if str(out.get("tipo", "")).upper() == "AUTOICA":
                out["autorretenciones"] = ws["AB35"].value
                out["total_impuesto_cargo"] = ws["AB35"].value
            out["ciiu"] = [str(ws.cell(r, 11).value) for r in range(24, 28) if isinstance(ws.cell(r, 11).value, (int, float))]
        else:
            # ReteICA cambia de geometría; se localizan renglones por su etiqueta y se toma el último número.
            for row in ws.iter_rows(max_row=min(ws.max_row, 100), max_col=min(ws.max_column, 80), values_only=True):
                label = " ".join(str(v) for v in row if isinstance(v, str)).upper()
                nums = [v for v in row if isinstance(v, (int, float))]
                if not nums: continue
                value = nums[-1]
                if "TOTAL RETENCIONES A TITULO" in label: out["retenciones"] = value
                if "TOTAL A PAGAR" in label: out["valor_pagar"] = value
        return out
    finally:
        wb.close()
