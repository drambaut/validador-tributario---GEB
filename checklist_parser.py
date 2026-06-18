from pathlib import Path
from openpyxl import load_workbook
from models import ChecklistRule


def parse_checklist(path: Path) -> list[ChecklistRule]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rules: list[ChecklistRule] = []
        for row in ws.iter_rows(min_row=1, max_col=11, values_only=True):
            number, concept, reference = row[1], row[2], row[9]
            if isinstance(number, (int, float)) and isinstance(concept, str):
                rules.append(ChecklistRule(int(number), concept.strip(), str(reference or "").strip()))
        return rules
    finally:
        wb.close()
